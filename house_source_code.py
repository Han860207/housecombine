#!/usr/bin/env python
# coding: utf-8

# In[1]:


county=['基隆市','台北市','新北市','桃園市','新竹市','新竹縣','苗栗縣',
        '台中市','彰化縣','南投縣','雲林縣','嘉義市','嘉義縣','台南市',
        '高雄市','屏東縣','台東縣','花蓮縣','宜蘭縣','澎湖縣','金門縣','連江縣'
       ]
area_data = {
    '台北市': [
        '中正區', '大同區', '中山區', '萬華區', '信義區', '松山區', '大安區', '南港區', '北投區', '內湖區', '士林區', '文山區'
    ],
    '新北市': [
        '板橋區', '新莊區', '泰山區', '林口區', '淡水區', '金山區', '八里區', '萬里區', '石門區', '三芝區', '瑞芳區', '汐止區', '平溪區', '貢寮區', '雙溪區', '深坑區', '石碇區', '新店區', '坪林區', '烏來區', '中和區', '永和區', '土城區', '三峽區', '樹林區', '鶯歌區', '三重區', '蘆洲區', '五股區'
    ],
    '基隆市': [
        '仁愛區', '中正區', '信義區', '中山區', '安樂區', '暖暖區', '七堵區'
    ],
    '桃園市': [
        '桃園區', '中壢區', '平鎮區', '八德區', '楊梅區', '蘆竹區', '龜山區', '龍潭區', '大溪區', '大園區', '觀音區', '新屋區', '復興區'
    ],
    '新竹縣': [
        '竹北市', '竹東鎮', '新埔鎮', '關西鎮', '峨眉鄉', '寶山鄉', '北埔鄉', '橫山鄉', '芎林鄉', '湖口鄉', '新豐鄉', '尖石鄉', '五峰鄉'
    ],
    '新竹市': [
        '東區', '北區', '香山區'
    ],
    '苗栗縣': [
        '苗栗市', '通霄鎮', '苑裡鎮', '竹南鎮', '頭份鎮', '後龍鎮', '卓蘭鎮', '西湖鄉', '頭屋鄉', '公館鄉', '銅鑼鄉', '三義鄉', '造橋鄉', '三灣鄉', '南庄鄉', '大湖鄉', '獅潭鄉', '泰安鄉'
    ],
    '台中市': [
        '中區', '東區', '南區', '西區', '北區', '北屯區', '西屯區', '南屯區', '太平區', '大里區', '霧峰區', '烏日區', '豐原區', '后里區', '東勢區', '石岡區', '新社區', '和平區', '神岡區', '潭子區', '大雅區', '大肚區', '龍井區', '沙鹿區', '梧棲區', '清水區', '大甲區', '外埔區', '大安區'
    ],
    '南投縣': [
        '南投市', '埔里鎮', '草屯鎮', '竹山鎮', '集集鎮', '名間鄉', '鹿谷鄉', '中寮鄉', '魚池鄉', '國姓鄉', '水里鄉', '信義鄉', '仁愛鄉'
    ],
    '彰化縣': [
        '彰化市', '員林鎮', '和美鎮', '鹿港鎮', '溪湖鎮', '二林鎮', '田中鎮', '北斗鎮', '花壇鄉', '芬園鄉', '大村鄉', '永靖鄉', '伸港鄉', '線西鄉', '福興鄉', '秀水鄉', '埔心鄉', '埔鹽鄉', '大城鄉', '芳苑鄉', '竹塘鄉', '社頭鄉', '二水鄉', '田尾鄉', '埤頭鄉', '溪州鄉'
    ],
    '雲林縣': [
        '斗六市', '斗南鎮', '虎尾鎮', '西螺鎮', '土庫鎮', '北港鎮', '莿桐鄉', '林內鄉', '古坑鄉', '大埤鄉', '崙背鄉', '二崙鄉', '麥寮鄉', '台西鄉', '東勢鄉', '褒忠鄉', '四湖鄉', '口湖鄉', '水林鄉', '元長鄉'
    ],
    '嘉義縣': [
        '太保市', '朴子市', '布袋鎮', '大林鎮', '民雄鄉', '溪口鄉', '新港鄉', '六腳鄉', '東石鄉', '義竹鄉', '鹿草鄉', '水上鄉', '中埔鄉', '竹崎鄉', '梅山鄉', '番路鄉', '大埔鄉', '阿里山鄉'
    ],
    '嘉義市': [
        '東區', '西區'
    ],
    '台南市': [
        '中西區', '東區', '南區', '北區', '安平區', '安南區', '永康區', '歸仁區', '新化區', '左鎮區', '玉井區', '楠西區', '南化區', '仁德區', '關廟區', '龍崎區', '官田區', '麻豆區', '佳里區', '西港區', '七股區', '將軍區', '學甲區', '北門區', '新營區', '後壁區', '白河區', '東山區', '六甲區', '下營區', '柳營區', '鹽水區', '善化區', '大內區', '山上區', '新市區', '安定區'
    ],
    '高雄市': [
        '楠梓區', '左營區', '鼓山區', '三民區', '鹽埕區', '前金區', '新興區', '苓雅區', '前鎮區', '小港區', '旗津區', '鳳山區', '大寮區', '鳥松區', '林園區', '仁武區', '大樹區', '大社區', '岡山區', '路竹區', '橋頭區', '梓官區', '彌陀區', '永安區', '燕巢區', '田寮區', '阿蓮區', '茄萣區', '湖內區', '旗山區', '美濃區', '內門區', '杉林區', '甲仙區', '六龜區', '茂林區', '桃源區', '那瑪夏區'
    ],
    '屏東縣': [
        '屏東市', '潮州鎮', '東港鎮', '恆春鎮', '萬丹鄉', '長治鄉', '麟洛鄉', '九如鄉', '里港鄉', '鹽埔鄉', '高樹鄉', '萬巒鄉', '內埔鄉', '竹田鄉', '新埤鄉', '枋寮鄉', '新園鄉', '崁頂鄉', '林邊鄉', '南州鄉', '佳冬鄉', '琉球鄉', '車城鄉', '滿州鄉', '枋山鄉', '霧台鄉', '瑪家鄉', '泰武鄉', '來義鄉', '春日鄉', '獅子鄉', '牡丹鄉', '三地門鄉'
    ],
    '宜蘭縣': [
        '宜蘭市', '羅東鎮', '蘇澳鎮', '頭城鎮', '礁溪鄉', '壯圍鄉', '員山鄉', '冬山鄉', '五結鄉', '三星鄉', '大同鄉', '南澳鄉'
    ],
    '花蓮縣': [
        '花蓮市', '鳳林鎮', '玉里鎮', '新城鄉', '吉安鄉', '壽豐鄉', '秀林鄉', '光復鄉', '豐濱鄉', '瑞穗鄉', '萬榮鄉', '富里鄉', '卓溪鄉'
    ],
    '台東縣': [
        '台東市', '成功鎮', '關山鎮', '長濱鄉', '海端鄉', '池上鄉', '東河鄉', '鹿野鄉', '延平鄉', '卑南鄉', '金峰鄉', '大武鄉', '達仁鄉', '綠島鄉', '蘭嶼鄉', '太麻里鄉'
    ],
    '澎湖縣': [
        '馬公市', '湖西鄉', '白沙鄉', '西嶼鄉', '望安鄉', '七美鄉'
    ],
    '金門縣': [
        '金城鎮', '金湖鎮', '金沙鎮', '金寧鄉', '烈嶼鄉', '烏坵鄉'
    ],
    '連江縣': [
        '南竿鄉', '北竿鄉', '莒光鄉', '東引鄉'
    ]
}
eng = {"基隆市" : "Keelung-city",
"新北市" : "NewTaipei-city",
"台北市" : "Taipei-city",
"桃園市" : "Taoyuan-city",
"新竹縣" : "Hsinchu-county",
"新竹市" : "Hsinchu-city",
"苗栗市" : "Miaoli-city",
"苗栗縣" : "Miaoli-county",
"台中市" : "Taichung-city",
"彰化縣" : "Changhua-county",
"彰化市" : "Changhua-city",
"南投市" : "Nantou-city",
"南投縣" : "Nantou-county",
"雲林縣" : "Yunlin-county",
"嘉義縣" : "Chiayi-county",
"嘉義市" : "Chiayi-city",
"台南市" : "Tainan-city",
"高雄市" : "Kaohsiung-city",
"屏東縣" : "Pingtung-county",
"屏東市" : "Pingtung-city",
"宜蘭縣" : "Yilan-county",
"宜蘭市" : "Yilan-city",
"花蓮縣" : "Hualien-county",
"花蓮市" : "Hualien-city",
"台東市" : "Taitung-city",
"台東縣" : "Taitung-county",
"澎湖縣" : "Penghu-county",
"綠島" : "Green-Island",
"蘭嶼" : "Orchid-Island",
"金門縣" : "Kinmen-county",
"連江縣" : "Lienchiang-county"}
num={"中正區": "100",
"大同區": "103",
"中山區": "104",
"松山區": "105",
"大安區": "106",
"萬華區": "108",
"信義區": "110",
"士林區": "111",
"北投區": "112",
"內湖區": "114",
"南港區": "115",
"文山區": "116",
"釣魚臺": "290",
"仁愛區": "200",
"信義區": "201",
"中正區": "202",
"中山區": "203",
"安樂區": "204",
"暖暖區": "205",
"七堵區": "206",
"萬里區": "207",
"金山區": "208",
"板橋區": "220",
"汐止區": "221",
"深坑區": "222",
"石碇區": "223",
"瑞芳區": "224",
"平溪區": "226",
"雙溪區": "227",
"貢寮區": "228",
"新店區": "231",
"坪林區": "232",
"烏來區": "233",
"永和區": "234",
"中和區": "235",
"土城區": "236",
"三峽區": "237",
"樹林區": "238",
"鶯歌區": "239",
"三重區": "241",
"新莊區": "242",
"泰山區": "243",
"林口區": "244",
"蘆洲區": "247",
"五股區": "248",
"八里區": "249",
"淡水區": "251",
"三芝區": "252",
"石門區": "253",
"宜蘭市": "260",
"頭城鎮": "261",
"礁溪鄉": "262",
"壯圍鄉": "263",
"員山鄉": "264",
"羅東鎮": "265",
"三星鄉": "266",
"大同鄉": "267",
"五結鄉": "268",
"冬山鄉": "269",
"蘇澳鎮": "270",
"南澳鄉": "272",
"南竿鄉": "209",
"北竿鄉": "210",
"莒光鄉": "211",
"東引鄉": "212",
"東區": "300",
"北區": "300",
"香山區": "300",
"竹北市": "302",
"湖口鄉": "303",
"新豐鄉": "304",
"新埔鎮": "305",
"關西鎮": "306",
"芎林鄉": "307",
"寶山鄉": "308",
"竹東鎮": "310",
"五峰鄉": "311",
"橫山鄉": "312",
"尖石鄉": "313",
"北埔鄉": "314",
"峨眉鄉": "315",
"中壢區": "320",
"平鎮區": "324",
"龍潭區": "325",
"楊梅區": "326",
"新屋區": "327",
"觀音區": "328",
"桃園區": "330",
"龜山區": "333",
"八德區": "334",
"大溪區": "335",
"復興區": "336",
"大園區": "337",
"蘆竹區": "338",
"竹南鎮": "350",
"頭份市": "351",
"三灣鄉": "352",
"南庄鄉": "353",
"獅潭鄉": "354",
"後龍鎮": "356",
"通霄鎮": "357",
"苑裡鎮": "358",
"苗栗市": "360",
"造橋鄉": "361",
"頭屋鄉": "362",
"公館鄉": "363",
"大湖鄉": "364",
"泰安鄉": "365",
"銅鑼鄉": "366",
"三義鄉": "367",
"西湖鄉": "368",
"卓蘭鎮": "369",
"中區": "400",
"東區": "401",
"南區": "402",
"西區": "403",
"北區": "404",
"北屯區": "406",
"西屯區": "407",
"南屯區": "408",
"太平區": "411",
"大里區": "412",
"霧峰區": "413",
"烏日區": "414",
"豐原區": "420",
"后里區": "421",
"石岡區": "422",
"東勢區": "423",
"和平區": "424",
"新社區": "426",
"潭子區": "427",
"大雅區": "428",
"神岡區": "429",
"大肚區": "432",
"沙鹿區": "433",
"龍井區": "434",
"梧棲區": "435",
"清水區": "436",
"大甲區": "437",
"外埔區": "438",
"大安區": "439",
"彰化市": "500",
"芬園鄉": "502",
"花壇鄉": "503",
"秀水鄉": "504",
"鹿港鎮": "505",
"福興鄉": "506",
"線西鄉": "507",
"和美鎮": "508",
"伸港鄉": "509",
"員林鎮": "510",
"社頭鄉": "511",
"永靖鄉": "512",
"埔心鄉": "513",
"溪湖鎮": "514",
"大村鄉": "515",
"埔鹽鄉": "516",
"田中鎮": "520",
"北斗鎮": "521",
"田尾鄉": "522",
"埤頭鄉": "523",
"溪州鄉": "524",
"竹塘鄉": "525",
"二林鎮": "526",
"大城鄉": "527",
"芳苑鄉": "528",
"二水鄉": "530",
"南投市": "540",
"中寮鄉": "541",
"草屯鎮": "542",
"國姓鄉": "544",
"埔里鎮": "545",
"仁愛鄉": "546",
"名間鄉": "551",
"集集鎮": "552",
"水里鄉": "553",
"魚池鄉": "555",
"信義鄉": "556",
"竹山鎮": "557",
"鹿谷鄉": "558",
"東區": "600",
"西區": "600",
"番路鄉": "602",
"梅山鄉": "603",
"竹崎鄉": "604",
"阿里山": "605",
"中埔鄉": "606",
"大埔鄉": "607",
"水上鄉": "608",
"鹿草鄉": "611",
"太保市": "612",
"朴子市": "613",
"東石鄉": "614",
"六腳鄉": "615",
"新港鄉": "616",
"民雄鄉": "621",
"大林鎮": "622",
"溪口鄉": "623",
"義竹鄉": "624",
"布袋鎮": "625",
"斗南鎮": "630",
"大埤鄉": "631",
"虎尾鎮": "632",
"土庫鎮": "633",
"褒忠鄉": "634",
"東勢鄉": "635",
"臺西鄉": "636",
"崙背鄉": "637",
"麥寮鄉": "638",
"斗六市": "640",
"林內鄉": "643",
"古坑鄉": "646",
"莿桐鄉": "647",
"西螺鎮": "648",
"二崙鄉": "649",
"北港鎮": "651",
"水林鄉": "652",
"口湖鄉": "653",
"四湖鄉": "654",
"元長鄉": "655",
"中西區": "700",
"東區": "701",
"南區": "702",
"北區": "704",
"安平區": "708",
"安南區": "709",
"永康區": "710",
"歸仁區": "711",
"新化區": "712",
"左鎮區": "713",
"玉井區": "714",
"楠西區": "715",
"南化區": "716",
"仁德區": "717",
"關廟區": "718",
"龍崎區": "719",
"官田區": "720",
"麻豆區": "721",
"佳里區": "722",
"西港區": "723",
"七股區": "724",
"將軍區": "725",
"學甲區": "726",
"北門區": "727",
"新營區": "730",
"後壁區": "731",
"白河區": "732",
"東山區": "733",
"六甲區": "734",
"下營區": "735",
"柳營區": "736",
"鹽水區": "737",
"善化區": "741",
"大內區": "742",
"山上區": "743",
"新市區": "744",
"安定區": "745",
"新興區": "800",
"前金區": "801",
"苓雅區": "802",
"鹽埕區": "803",
"鼓山區": "804",
"旗津區": "805",
"前鎮區": "806",
"三民區": "807",
"楠梓區": "811",
"小港區": "812",
"左營區": "813",
"仁武區": "814",
"大社區": "815",
"岡山區": "820",
"路竹區": "821",
"阿蓮區": "822",
"田寮區": "823",
"燕巢區": "824",
"橋頭區": "825",
"梓官區": "826",
"彌陀區": "827",
"永安區": "828",
"湖內區": "829",
"鳳山區": "830",
"大寮區": "831",
"林園區": "832",
"鳥松區": "833",
"大樹區": "840",
"旗山區": "842",
"美濃區": "843",
"六龜區": "844",
"內門區": "845",
"杉林區": "846",
"甲仙區": "847",
"桃源區": "848",
"那瑪夏": "849",
"茂林區": "851",
"茄萣區": "852",
"屏東市": "900",
"三地門": "901",
"霧臺鄉": "902",
"瑪家鄉": "903",
"九如鄉": "904",
"里港鄉": "905",
"高樹鄉": "906",
"鹽埔鄉": "907",
"長治鄉": "908",
"麟洛鄉": "909",
"竹田鄉": "911",
"內埔鄉": "912",
"萬丹鄉": "913",
"潮州鎮": "920",
"泰武鄉": "921",
"來義鄉": "922",
"萬巒鄉": "923",
"崁頂鄉": "924",
"新埤鄉": "925",
"南州鄉": "926",
"林邊鄉": "927",
"東港鎮": "928",
"琉球鄉": "929",
"佳冬鄉": "931",
"新園鄉": "932",
"枋寮鄉": "940",
"枋山鄉": "941",
"春日鄉": "942",
"獅子鄉": "943",
"車城鄉": "944",
"牡丹鄉": "945",
"恆春鎮": "946",
"滿州鄉": "947",
"馬公市": "880",
"西嶼鄉": "881",
"望安鄉": "882",
"七美鄉": "883",
"白沙鄉": "884",
"湖西鄉": "885",
"金沙鎮": "890",
"金湖鎮": "891",
"金寧鄉": "892",
"金城鎮": "893",
"烈嶼鄉": "894",
"烏坵鄉": "896",
"臺東市": "950",
"綠島鄉": "951",
"蘭嶼鄉": "952",
"延平鄉": "953",
"卑南鄉": "954",
"鹿野鄉": "955",
"關山鎮": "956",
"海端鄉": "957",
"池上鄉": "958",
"東河鄉": "959",
"成功鎮": "961",
"長濱鄉": "962",
"太麻里": "963",
"金峰鄉": "964",
"大武鄉": "965",
"達仁鄉": "966",
"花蓮市": "970",
"新城鄉": "971",
"秀林鄉": "972",
"吉安鄉": "973",
"壽豐鄉": "974",
"鳳林鎮": "975",
"光復鄉": "976",
"豐濱鄉": "977",
"瑞穗鄉": "978",
"萬榮鄉": "979",
"玉里鎮": "981",
"卓溪鄉": "982",
"富里鄉": "983"}

import tkinter as tk
from tkinter import ttk
###
###視窗設計
###
app = tk.Tk() 
app.geometry('1500x380')
app.config(bg='#ADADAD')
app.resizable(0,0)
app.title('買屋查詢')
##
label =tk.Label(app,text ='房仲',bg='#ADADAD')
label.grid(column = 0 ,row =0 ,sticky = 'w')
combo = ttk.Combobox(app,values =['永慶','信義'], width =8)
combo.grid(column = 0,row=1,sticky = 'w')


labelTop = tk.Label(app,text = '縣市',bg='#ADADAD')
labelTop.grid(column=1, row=0,sticky = 'w')
comboExample = ttk.Combobox(app,values=county,width =13)
comboExample.grid(column=1, row=1,sticky = 'w')

labelTop1 = tk.Label(app,text = '鄉鎮市區',bg='#ADADAD')
labelTop1.grid(column=2, row=0,sticky = 'w' )

###當縣市被選取時,更改鄉鎮市區的欄位
def go (self):
    global comboExample
    global comboExample1
    area = comboExample.get()
    comboExample1 = ttk.Combobox(app,values=area_data[area],width =13)
    comboExample1.grid(column=2, row=1,sticky = 'w')
    comboExample1.current(1)
comboExample1 = ttk.Combobox(app,width =13)
comboExample1.grid(column=2, row=1,sticky = 'w')
comboExample.bind("<<ComboboxSelected>>",go)

### 搭配crawl_insert使用,爬永慶的資料
def crawl(page):
    global county1 , region
    county1 = comboExample.get()
    region = comboExample1.get()
    import requests
    from bs4 import BeautifulSoup
    head={'User-Agent':'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36'}
    res=requests.get("https://buy.yungching.com.tw/region/"+county1+'-'+region+'_c?pg='+str(page),headers=head)
    soup=BeautifulSoup(res.text)
    t = soup.find_all(class_="m-list-item")
    return soup,t

def crawl_insert():
    from tkinter import messagebox
    import time
    for i in tree.get_children(): 
        tree.delete(i)
    cnt = [1,2,3,4,5]
    count=0
    page = 0
    soup ,t = crawl(page)
    while soup.find_all('a',ga_label="buy_page_next")[0]['href'] != '':
        page+=1
        cnt.append(soup.find_all('a',ga_label="buy_page_next")[0]['href'])
        cnt.pop(0)
        if cnt[-1] == cnt[-5]:
            break
        soup ,t = crawl(page)
        for x in t:
            tree.insert("",tk.END,text='' ,values=(
            x.div.a.h3.string,
            x.div.div.span.string,
            x.select('ul')[0].select('li')[0].string,
            'https://buy.yungching.com.tw/'+x.a['href'],
            x.find(class_='item-info-detail').select('li')[5].text,
            x.select('ul')[0].select('li')[4].string,
            x.find(class_='price-num').string +'萬'
            ))
            count+=1
        time.sleep(7)
    else :
        page+=1
        soup ,t = crawl(page)
        for x in t:
            tree.insert("",tk.END,text='' ,values=(
            x.div.a.h3.string,
            x.div.div.span.string,
            x.select('ul')[0].select('li')[0].string,
            'https://buy.yungching.com.tw/'+x.a['href'],
            x.find(class_='item-info-detail').select('li')[5].text,
            x.select('ul')[0].select('li')[4].string,
            x.find(class_='price-num').string +'萬'
            ))
            count+=1
    messagebox.showinfo('Notice','End of search. 共%s筆資料'%count)
###爬信義的資料
def crawl_xinyi():
    from tkinter import messagebox
    global countyEng,region,eng,num
    region = comboExample1.get()
    countyEng =eng[comboExample.get()]
    num = num[region]
    for i in tree.get_children(): 
        tree.delete(i)
    import requests
    import time
    from bs4 import BeautifulSoup
    count = 0 
    page = 0
    head={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36 Edg/96.0.1054.29'}
    res=requests.get("https://www.sinyi.com.tw/buy/list/%s/%s-zip/Taipei-R-mrtline/03-mrt/default-desc/"%(countyEng,num)+str(page),headers=head)
    soup=BeautifulSoup(res.text)
    while page < int(soup.find_all('li')[-2].a.text):
        page+=1
        res=requests.get("https://www.sinyi.com.tw/buy/list/%s/%s-zip/Taipei-R-mrtline/03-mrt/default-desc/"%(countyEng,num)+str(page),headers=head)
        soup=BeautifulSoup(res.text)
        t = soup.find_all(class_='buy-list-item')
        for x in t :
            url = 'https://www.sinyi.com.tw'+x.a['href']
            res2 = requests.get(url)
            soup2 = BeautifulSoup(res2.text)
            tree.insert("",tk.END,text='' ,values=(
            x.find_all(class_='LongInfoCard_Type_Name')[0].text,
            x.find_all(class_='LongInfoCard_Type_Address')[0].select('span')[0].text,
            x.find_all(class_='LongInfoCard_Type_Address')[0].select('span')[2].text,
            url,
            x.find_all(class_='LongInfoCard_Type_HouseInfo')[0].select('span')[0].text,
            x.find_all(class_='LongInfoCard_Type_HouseInfo')[0].select('span')[1].text,
            soup2.find_all(class_='buy-content-title-total-price')[0].text))
            count+=1
        time.sleep(7)
    messagebox.showinfo('Notice','End of search. 共%s筆資料'%count)
### 設置查詢鈕
btn = tk.Button(app,text= '查詢')
###當房仲被選取時,設定查詢鈕該使用永慶的爬蟲還是信義的爬蟲          
def one(self):
    global combo_one
    combo_one = combo.get()
    if combo_one =='信義':
        btn.config(command = crawl_xinyi)
    else :
        btn.config(command = crawl_insert)
combo.bind('<<ComboboxSelected>>',one)
btn.config()
btn.grid(column = 3 , row = 1,sticky = 'w')


###下載鈕
def download():
    region = comboExample1.get()
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '標題'
    ws['B1'] = '售屋地址'
    ws['C1'] = '售屋類型'
    ws['D1'] = '網址'
    ws['E1'] = '建坪'
    ws['F1'] = '建物坪數'
    ws['G1'] = '價格'
    for i in tree.get_children(): 
        ws.append(tree.item(i)['values'])
    wb.save(county1+'_'+region+'.xlsx')

btn_download = tk.Button(app, text= '下載')
btn_download.grid(column = 7 , row = 1)
btn_download.config(command=download)


### 表格右方的下拉bar
scroll = tk.Scrollbar()
###表格設定
tree = ttk.Treeview(app,height=14,yscrollcommand = scroll.set)
tree['show'] = 'headings'
tree['columns'] = (1,2,3,4,5,6,7)
for i in range(1,8):
    tree.column(i,width=130)
tree.heading(1,text="標題")
tree.heading(2,text="售屋地址")
tree.heading(3,text="售屋類型")
tree.heading(4,text="網址")
tree.heading(5,text="建坪")
tree.heading(6,text="建物坪數")
tree.heading(7,text='價格')
tree.grid(column = 0,columnspan = 6,row = 2,rowspan = 2)



scroll.config(command=tree.yview)
scroll.grid(column = 6,row = 2,sticky='n'+'s'+'w',rowspan=2)

### 當點選同樣的資料兩下可以移動到該網站
def treeviewClick(self):
    import webbrowser
    curitem = tree.focus()
    webbrowser.open(tree.item(curitem)['values'][3])
tree.bind('<Double-1>', treeviewClick)

### 圖片用
import requests
import io
from PIL import Image ,ImageTk


image_bytes = requests.get('https://www.python.org/static/community_logos/python-logo.png').content
data_stream = io.BytesIO(image_bytes)
pil_image = Image.open(data_stream)
tk_image = ImageTk.PhotoImage(pil_image)
pic_label1 = tk.Label(app,image = tk_image , bg ='#ADADAD')
pic_label1.grid(row= 3,column = 8 ,sticky = 'n',rowspan =2 )

image_bytes2 = requests.get('https://chientinglu.files.wordpress.com/2020/11/666-1.png').content
data_stream2 = io.BytesIO(image_bytes2)
pil_image2 = Image.open(data_stream2)
tk_image2 = ImageTk.PhotoImage(pil_image2)
pic_label2 = tk.Label(app,image = tk_image2 , bg ='#ADADAD')
pic_label2.grid(row= 2,column = 8 ,sticky = 'w'+'n')
### end
app.mainloop()

